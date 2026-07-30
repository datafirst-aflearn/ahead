"""Build variable-availability-matrix.xlsx from EGRA_Harmonization_Waves_Details.xlsx."""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent
EXCEL_PATH = ROOT / "EGRA_Harmonization_Waves_Details.xlsx"
OUT_PATH = ROOT / "variable-availability-matrix.xlsx"

MASTER_TO_TASK = {
    "num_id": "num_id",
    "quant_comp": "quant_comp",
    "add": "addlvl1",
    "addlvl2": "addlvl2",
    "sub": "sublvl1",
    "sublvl2": "sublvl2",
    "miss_num": "miss_num",
    "word_prob": "word_prob",
    "list_comp": "list_comp",
    "l2_list_comp": "list_comp",
    "letter_name": "letter",
    "l2_letter_name": "letter",
    "letter_sound": "letter_sound",
    "l2_letter_sound": "letter_sound",
    "invent_word": "invent_word",
    "l2_invent_word": "invent_word",
    "oral_read": "oral_read",
    "oral_readB": "oral_read",
    "l2_oral_read": "oral_read",
    "read_comp": "read_comp",
    "read_compB": "read_comp",
    "l2_read_comp": "read_comp",
    "fam_word": "fam_word",
    "l2_fam_word": "fam_word",
    "mazeA": "maze",
    "mazeB": "maze",
    "pa_init_sound": "pa_init_sound",
    "l2_pa_init_sound": "pa_init_sound",
    "pa_df_init_snd": "pa_df_init_snd",
    "pa_df_fnl_snd": "pa_df_fnl_snd",
    "pa_num_sound": "pa_num_sound",
    "pa_phon_sound": "pa_num_sound",
    "phoneme_seg_a_en": "pa_num_sound",
    "l2_pa_num_sound": "pa_num_sound",
    "syllable_sound": "syllable_sound",
    "l2_syllable_sound": "syllable_sound",
    "dict": "dict_let",
    "dict_word": "word_dict",
    "dict_sent": "word_dict",
    "vocabA": "vocab",
    "vocabB": "vocab",
    "vocabC": "vocab",
    "l2_vocabA": "vocab",
    "l2_vocabB": "vocab",
    "l2_vocabC": "vocab",
    "l2_vocab_word": "vocab",
    "l2_oral_vocab": "oral_vocab",
}


def clean_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def load_task_meaning(wb):
    rows = list(wb["task meaning"].iter_rows(values_only=True))
    header = [clean_text(h) for h in rows[0]]
    tasks = []
    seen_prefix = set()
    for row in rows[1:]:
        if not any(row):
            continue
        rec = {header[i]: clean_text(row[i]) if i < len(row) else "" for i in range(len(header))}
        prefix = clean_text(rec.get("Subtask prefix", ""))
        if not prefix or prefix in seen_prefix:
            continue
        seen_prefix.add(prefix)
        tasks.append(
            {
                "Assessment": rec.get("Assessment", ""),
                "Task": rec.get("Task", ""),
                "Core": rec.get("Core", ""),
                "Subtask prefix": prefix,
                "Alternate prefixes found in datasets": rec.get(
                    "Alternate prefixes found in datasets", ""
                ),
            }
        )
    return tasks


def load_subtasks(wb):
    rows = list(wb["sub-tasks"].iter_rows(values_only=True))
    header = [clean_text(h) for h in rows[0]]
    study_cols = header[2:]
    present = {col: set() for col in study_cols}
    for row in rows[1:]:
        if not row or len(row) < 3:
            continue
        master = clean_text(row[1])
        task_id = MASTER_TO_TASK.get(master)
        if not task_id:
            continue
        for idx, col in enumerate(study_cols, start=2):
            val = row[idx] if idx < len(row) else None
            if val not in (None, ""):
                present[col].add(task_id)
    return study_cols, present, tasks_from_rows(rows, study_cols)


def tasks_from_rows(rows, study_cols):
    del rows, study_cols
    return None


def main():
    wb_in = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    tasks = load_task_meaning(wb_in)
    study_cols, present, _ = load_subtasks(wb_in)
    wb_in.close()

    wb_out = Workbook()
    desc_ws = wb_out.active
    desc_ws.title = "sub-tasks-desc"
    desc_header = [
        "Assessment",
        "Task",
        "Core",
        "Subtask prefix",
        "Alternate prefixes found in datasets",
    ]
    desc_ws.append(desc_header)
    for task in tasks:
        desc_ws.append([task[h] for h in desc_header])

    matrix_ws = wb_out.create_sheet("va-matrix")
    meta_cols = desc_header
    matrix_ws.append(meta_cols + study_cols)
    for task in tasks:
        prefix = task["Subtask prefix"]
        row = [task[h] for h in meta_cols]
        for col in study_cols:
            row.append("Yes" if prefix in present[col] else "No")
        matrix_ws.append(row)

    wb_out.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"  Tasks: {len(tasks)}")
    print(f"  Study columns: {len(study_cols)}")


if __name__ == "__main__":
    main()
